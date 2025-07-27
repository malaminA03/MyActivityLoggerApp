import tkinter as tk
from tkinter import ttk, messagebox, font as tkfont, filedialog
import time
import json
import os
import platform
import threading
import queue
import csv
import webbrowser
from datetime import datetime, date, timedelta
from collections import defaultdict

# Attempt to import necessary libraries
try:
    from pystray import MenuItem as item, Icon as icon
    from PIL import Image, ImageDraw
    TRAY_ENABLED = True
except ImportError:
    TRAY_ENABLED = False

try:
    from tkcalendar import DateEntry
    CALENDAR_ENABLED = True
except ImportError:
    CALENDAR_ENABLED = False

try:
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_ENABLED = True
except ImportError:
    MATPLOTLIB_ENABLED = False

try:
    import google.generativeai as genai
    AI_ENABLED = True
except ImportError:
    AI_ENABLED = False

try:
    from pynput import mouse, keyboard
    IDLE_DETECTION_ENABLED = True
except ImportError:
    IDLE_DETECTION_ENABLED = False

try:
    import pyperclip
    CLIPBOARD_ENABLED = True
except ImportError:
    CLIPBOARD_ENABLED = False

try:
    import openpyxl
    EXCEL_ENABLED = True
except ImportError:
    EXCEL_ENABLED = False

# --- Global variables ---
event_queue = queue.Queue()
last_activity_time = time.time()

# --- Listener Functions ---
def on_activity(*args):
    global last_activity_time
    last_activity_time = time.time()

def on_click(x, y, button, pressed):
    if pressed:
        event_queue.put(('click', f"Mouse clicked: {button}"))
        on_activity()

def start_listeners():
    if not IDLE_DETECTION_ENABLED: return
    mouse_listener = mouse.Listener(on_click=on_click, on_move=on_activity, on_scroll=on_activity)
    keyboard_listener = keyboard.Listener(on_press=on_activity)
    mouse_listener.start(); keyboard_listener.start(); mouse_listener.join(); keyboard_listener.join()

# --- Active Window Title Fetching Function ---
def get_active_window_title():
    system = platform.system()
    if system == 'Windows':
        try:
            import win32gui
            return win32gui.GetWindowText(win32gui.GetForegroundWindow())
        except ImportError: return "win32gui not installed"
    elif system == 'Linux':
        try:
            import subprocess
            root = subprocess.check_output(['xprop', '-root', '_NET_ACTIVE_WINDOW'])
            window_id = root.decode().strip().split()[-1]
            window_name = subprocess.check_output(['xprop', '-id', window_id, 'WM_NAME'])
            return window_name.decode().split('"', 1)[1].rsplit('"', 1)[0]
        except: return "Could not get window title"
    else: return f"Unsupported OS: {system}"

# --- Main Application Class ---
class ActivityLoggerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("AI Productivity Coach")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)

        # --- Load Config ---
        self.config_file = os.path.expanduser('~/.activity_logger_config.json')
        self.load_config()

        # --- Fonts ---
        self.primary_font = ("Segoe UI", 10)
        self.header_font = ("Segoe UI", 11)
        self.card_title_font = ("Segoe UI", 12, "bold")
        self.card_value_font = ("Segoe UI", 22, "bold")
        self.sidebar_title_font = ("Segoe UI", 16, "bold")
        self.button_font = ("Segoe UI", 10, "bold")
        self.link_font = ("Segoe UI", 12, "underline")

        # --- Main Layout ---
        self.sidebar_frame = tk.Frame(root, width=220)
        self.sidebar_frame.pack(side="left", fill="y")
        self.sidebar_frame.pack_propagate(False)
        
        container = tk.Frame(root)
        container.pack(side="left", fill="both", expand=True)

        self.main_frame = tk.Frame(container)
        self.main_frame.pack(fill="both", expand=True)
        
        # --- UI Design & Colors ---
        self.setup_theme()

        # --- Data tracking variables ---
        self.txt_logfile = os.path.expanduser('~/.activity_log.txt')
        self.excel_logfile = os.path.expanduser('~/.activity_log.xlsx')
        self.data = self.load_log_from_file()
        self.active_time_seconds, self.idle_time_seconds = self.pre_calculate_today_stats()
        self.app_usage = defaultdict(float)
        self.last_app = None
        self.last_app_start_time = time.time()

        # --- Create UI Elements ---
        self.create_sidebar()
        self.pages = {
            "Dashboard": DashboardPage(self.main_frame, self),
            "Logs": LogsPage(self.main_frame, self),
            "Reports": ReportsPage(self.main_frame, self),
            "Settings": SettingsPage(self.main_frame, self),
            "About": AboutPage(self.main_frame, self) # New About Page
        }
        self.create_ai_section(container)
        
        # --- App Startup ---
        self.last_logged_event = None
        self.is_idle = False
        self.running = True
        self.start_background_tasks()
        self.show_page("Dashboard")
        self.update_dashboard_live()
        self.root.protocol("WM_DELETE_WINDOW", self.hide_window)

    def create_image(self, width, height, color1, color2):
        image = Image.new('RGB', (width, height), color1)
        dc = ImageDraw.Draw(image)
        dc.rectangle((width // 2, 0, width, height // 2), fill=color2)
        dc.rectangle((0, height // 2, width // 2, height), fill=color2)
        return image

    def quit_window(self, icon=None, item=None):
        self.running = False
        self.update_app_usage()
        self.save_config()
        if icon:
            icon.stop()
        self.root.destroy()

    def show_window(self, icon, item):
        if icon:
            icon.stop()
        self.root.after(0, self.root.deiconify)

    def hide_window(self):
        self.root.withdraw()
        if TRAY_ENABLED:
            image = self.create_image(64, 64, 'black', 'white')
            menu = (item('Show', self.show_window), item('Quit', self.quit_window))
            tray_icon = icon('name', image, "AI Productivity Coach", menu)
            threading.Thread(target=tray_icon.run, daemon=True).start()
        else:
            print("pystray is not installed. Application will close.")
            self.quit_window()

    def setup_theme(self):
        theme_name = self.config.get('theme', 'WordPress Light')
        themes = {
            "WordPress Light": {"bg": "#F0F0F0", "frame": "#FFFFFF", "sidebar": "#23282D", "text": "#333333", "sidebar_text": "#EEEEEE", "accent": "#0073AA", "header_bg": "#E5E5E5"},
            "Dark Teal": {"bg": "#121212", "frame": "#1F1F1F", "sidebar": "#1F1F1F", "text": "#E0E0E0", "sidebar_text": "#E0E0E0", "accent": "#00ADB5", "header_bg": "#1F1F1F"},
            "Light Mint": {"bg": "#F0FFF0", "frame": "#FFFFFF", "sidebar": "#E0EEE0", "text": "#000000", "sidebar_text": "#000000", "accent": "#2E8B57", "header_bg": "#E0EEE0"},
        }
        self.theme_colors = themes.get(theme_name, themes["WordPress Light"])
        self.root.configure(bg=self.theme_colors["bg"])
        self.sidebar_frame.configure(bg=self.theme_colors["sidebar"])
        self.main_frame.configure(bg=self.theme_colors["bg"])
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook", background=self.theme_colors["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", font=self.header_font, padding=[10, 5], background=self.theme_colors["header_bg"], foreground=self.theme_colors["text"])
        style.map("TNotebook.Tab", background=[("selected", self.theme_colors["frame"])])
        style.configure("Treeview", background=self.theme_colors["frame"], foreground=self.theme_colors["text"], rowheight=28, fieldbackground=self.theme_colors["frame"], font=self.primary_font, borderwidth=0)
        style.map('Treeview', background=[('selected', self.theme_colors["accent"])])
        style.configure("Treeview.Heading", background=self.theme_colors["header_bg"], foreground=self.theme_colors["text"], font=self.header_font, relief="flat", padding=5)
        style.map("Treeview.Heading", background=[('active', self.theme_colors["accent"])])
    
    def create_sidebar(self):
        for widget in self.sidebar_frame.winfo_children(): widget.destroy()
        tk.Label(self.sidebar_frame, text="Activity Logger", font=self.sidebar_title_font, bg=self.theme_colors["sidebar"], fg=self.theme_colors["accent"]).pack(pady=20, padx=10)
        self.sidebar_buttons = {}
        self.create_sidebar_button("Dashboard"); self.create_sidebar_button("Logs");
        self.create_sidebar_button("Reports"); self.create_sidebar_button("Settings");
        self.create_sidebar_button("About") # New About button
    
    def create_sidebar_button(self, text):
        button = tk.Button(self.sidebar_frame, text=text, command=lambda t=text: self.show_page(t), font=self.header_font, bg=self.theme_colors["sidebar"], fg=self.theme_colors["sidebar_text"], relief="flat", anchor="w", padx=20, pady=10, activebackground=self.theme_colors["accent"], activeforeground=self.theme_colors["sidebar_text"])
        button.pack(fill="x")
        self.sidebar_buttons[text] = button
    
    def create_ai_section(self, parent):
        ai_frame = tk.LabelFrame(parent, text="AI Assistant", bg=self.theme_colors["bg"], fg=self.theme_colors["text"], font=self.card_title_font, padx=10, pady=10, borderwidth=0, highlightthickness=0)
        ai_frame.pack(fill="x", padx=20, pady=(0, 20))
        self.ai_response_var = tk.StringVar(value="> AI: I am ready to help you.")
        tk.Label(ai_frame, textvariable=self.ai_response_var, anchor='w', justify='left', wraplength=750, bg=self.theme_colors["bg"], fg=self.theme_colors["text"], font=self.primary_font).pack(fill='x', pady=5)
        query_frame = tk.Frame(ai_frame, bg=self.theme_colors["bg"])
        query_frame.pack(fill='x', pady=5)
        self.query_entry = tk.Entry(query_frame, font=self.primary_font, bg=self.theme_colors["frame"], fg=self.theme_colors["text"], relief='solid', borderwidth=1, insertbackground=self.theme_colors["text"])
        self.query_entry.pack(side='left', fill='x', expand=True, ipady=8, padx=(0,10))
        self.query_entry.bind("<Return>", self.handle_enter_key)
        tk.Button(query_frame, text="Ask", command=self.ask_ai_assistant, bg=self.theme_colors["accent"], fg="#FFFFFF", font=self.button_font, relief='flat', padx=15, pady=5).pack(side='right')
    
    def show_page(self, page_name):
        for name, page in self.pages.items(): page.pack_forget()
        for name, button in self.sidebar_buttons.items(): button.config(bg=self.theme_colors["sidebar"])
        self.pages[page_name].pack(fill="both", expand=True, padx=20, pady=20)
        self.sidebar_buttons[page_name].config(bg=self.theme_colors["accent"])
        self.pages[page_name].on_show()
    
    def load_config(self):
        try:
            with open(self.config_file, 'r') as f: self.config = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.config = {'theme': 'WordPress Light', 'idle_threshold_minutes': 5, 'monitor_clicks': True, 'monitor_clipboard': True, 'daily_work_goal_hours': 4, 'work_keywords': 'code,develop,github,stackoverflow,docs,visual studio', 'entertainment_keywords': 'youtube,facebook,netflix,game,movie,music', 'break_reminder_enabled': True, 'break_reminder_interval_minutes': 60}
            self.save_config()
    
    def save_config(self):
        with open(self.config_file, 'w') as f: json.dump(self.config, f, indent=4)
    
    def start_background_tasks(self):
        threading.Thread(target=self.track_activity, daemon=True).start()
        if self.config.get('monitor_clipboard', True) and CLIPBOARD_ENABLED: 
            threading.Thread(target=self.track_clipboard, daemon=True).start()
        if self.config.get('break_reminder_enabled', True):
            threading.Thread(target=self.track_break_reminder, daemon=True).start()
        if IDLE_DETECTION_ENABLED:
            threading.Thread(target=start_listeners, daemon=True).start()
        self.root.after(100, self.process_queue)
    
    def track_clipboard(self):
        last_content = ""
        while self.running:
            if not self.config.get('monitor_clipboard', True): time.sleep(5); continue
            try:
                current_content = pyperclip.paste()
                if current_content and current_content != last_content:
                    last_content = current_content
                    log_content = (current_content[:100] + '...') if len(current_content) > 100 else current_content
                    event_queue.put(('clipboard', f'"{log_content}"'))
            except Exception: last_content = ""
            time.sleep(2)
            
    def track_break_reminder(self):
        while self.running:
            interval_minutes = self.config.get('break_reminder_interval_minutes', 60)
            time.sleep(interval_minutes * 60)
            if self.running and self.config.get('break_reminder_enabled', True) and not self.is_idle:
                messagebox.showinfo("Break Time!", f"You've been working for {interval_minutes} minutes. Time for a short break!")
    
    def process_queue(self):
        try:
            while not event_queue.empty():
                event_type, event_description = event_queue.get_nowait()
                if event_type == 'click' and not self.config.get('monitor_clicks', True): continue
                self.log_event(event_type, event_description)
        finally:
            self.root.after(100, self.process_queue)
    
    def track_activity(self):
        while self.running:
            time.sleep(1)
            time_since_last_activity = time.time() - last_activity_time
            if time_since_last_activity > self.config.get('idle_threshold_minutes', 5) * 60:
                if not self.is_idle:
                    self.is_idle = True
                    self.log_event('activity', "System Event: User is Idle")
                    self.update_app_usage()
                self.idle_time_seconds += 1
            else:
                if self.is_idle:
                    self.is_idle = False
                    self.log_event('activity', "System Event: User is Active")
                    self.last_app_start_time = time.time()
                self.active_time_seconds += 1
                title = get_active_window_title()
                if title and f"Window: {title}" != self.last_logged_event:
                    self.log_event('activity', f"Window: {title}")
                    self.update_app_usage(title)
    
    def update_app_usage(self, new_app=None):
        now = time.time()
        if self.last_app:
            duration = now - self.last_app_start_time
            self.app_usage[self.last_app] += duration
        self.last_app = new_app
        self.last_app_start_time = now
    
    def update_dashboard_live(self):
        if self.pages["Dashboard"].winfo_exists(): self.pages["Dashboard"].update_all_stats()
        self.root.after(5000, self.update_dashboard_live)
    
    def categorize_app(self, window_title):
        title = window_title.lower()
        work_keys = [k.strip() for k in self.config.get('work_keywords', '').lower().split(',')]
        ent_keys = [k.strip() for k in self.config.get('entertainment_keywords', '').lower().split(',')]
        if any(key in title for key in work_keys if key): return "Work"
        if any(key in title for key in ent_keys if key): return "Entertainment"
        return "Other"
    
    def log_event(self, event_type, event_description):
        if event_description == self.last_logged_event and event_type == 'activity': return
        if event_type == 'activity': self.last_logged_event = event_description
        timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
        entry = {'time': timestamp, 'type': event_type, 'event': event_description}
        self.data.append(entry)
        self.append_to_txt_log(f"{timestamp} - [{event_type.upper()}] {event_description}")
        self.append_to_excel_log([timestamp, event_type.upper(), event_description])
        if self.pages["Logs"].winfo_exists(): self.pages["Logs"].add_log(event_type, timestamp, event_description)
    
    def load_log_from_file(self):
        if not os.path.exists(self.txt_logfile): return []
        data = []
        with open(self.txt_logfile, 'r', encoding='utf-8') as f:
            for line in f:
                try:
                    parts = line.strip().split(' - ', 2)
                    data.append({'time': parts[0], 'type': parts[1][1:-1].lower(), 'event': parts[2]})
                except: continue
        return data
    
    def pre_calculate_today_stats(self):
        today_str = date.today().strftime('%Y-%m-%d')
        todays_logs = [entry for entry in self.data if entry.get('time', '').startswith(today_str)]
        active_s, idle_s = 0, 0
        for i in range(len(todays_logs) - 1):
            try:
                t1 = datetime.strptime(todays_logs[i]['time'], '%Y-%m-%d %H:%M:%S')
                t2 = datetime.strptime(todays_logs[i+1]['time'], '%Y-%m-%d %H:%M:%S')
                duration = (t2 - t1).total_seconds()
                if duration > 1800: continue
                if "User is Idle" in todays_logs[i].get('event', ''): idle_s += duration
                else: active_s += duration
            except: continue
        return active_s, idle_s
    
    def append_to_txt_log(self, log_line):
        try:
            with open(self.txt_logfile, 'a', encoding='utf-8') as f: f.write(log_line + '\n')
        except IOError as e: print(f"Error writing to TXT log: {e}")
    
    def append_to_excel_log(self, row_data):
        if not EXCEL_ENABLED: return
        try:
            workbook = openpyxl.load_workbook(self.excel_logfile) if os.path.exists(self.excel_logfile) else openpyxl.Workbook()
            sheet = workbook.active
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1,1).value is None: sheet.append(["Timestamp", "Type", "Event"])
            sheet.append(row_data)
            workbook.save(self.excel_logfile)
        except Exception as e: print(f"Error writing to Excel log: {e}")
    
    def handle_enter_key(self, event): self.ask_ai_assistant()
    
    def ask_ai_assistant(self):
        user_question = self.query_entry.get()
        if not user_question: return
        self.query_entry.delete(0, tk.END)
        self.ai_response_var.set(f"> User: {user_question}\n> AI: Thinking...")
        threading.Thread(target=self.get_ai_response_from_api, args=(user_question,)).start()
    
    def get_ai_response_from_api(self, user_question):
        try:
            API_KEY = "YOUR_GEMINI_API_KEY_HERE" # Replace with your actual key
            genai.configure(api_key=API_KEY)
            model = genai.GenerativeModel('gemini-1.5-flash')
            with open(self.txt_logfile, 'r', encoding='utf-8') as f: log_data = f.read()
            prompt = f"Analyze the following log data and answer the user's question concisely in English.\n\nLog Data:\n---\n{log_data}\n---\n\nUser's Question: \"{user_question}\""
            response = model.generate_content(prompt)
            self.ai_response_var.set(f"> User: {user_question}\n> AI: {response.text}")
        except Exception as e: self.ai_response_var.set(f"> AI Error: An error occurred.\nDetails: {str(e)}")
    
    def on_close(self):
        self.running = False
        self.update_app_usage()
        self.save_config()
        self.root.destroy()

# --- Page Classes ---
class BasePage(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg=controller.theme_colors["bg"])
        self.controller = controller
    def on_show(self): pass

class DashboardPage(BasePage):
    def __init__(self, parent, controller):
        super().__init__(parent, controller)
        cards_frame = tk.Frame(self, bg=self.controller.theme_colors["bg"])
        cards_frame.pack(fill="x", pady=(0, 20))
        self.active_time_var = tk.StringVar(value="--:--"); self.idle_time_var = tk.StringVar(value="--:--")
        self.most_used_var = tk.StringVar(value="N/A"); self.productivity_var = tk.StringVar(value="--%")
        self.create_card(cards_frame, "Active Time", self.active_time_var); self.create_card(cards_frame, "Idle Time", self.idle_time_var)
        self.create_card(cards_frame, "Most Used App", self.most_used_var); self.create_card(cards_frame, "Productivity", self.productivity_var)
        goal_frame = tk.LabelFrame(self, text="Daily Goal", bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], font=self.controller.card_title_font, padx=10, pady=10, borderwidth=0)
        goal_frame.pack(fill="x", pady=(0, 20))
        self.goal_progress = ttk.Progressbar(goal_frame, orient="horizontal", length=300, mode="determinate")
        self.goal_progress.pack(fill="x", pady=5)
        self.goal_label_var = tk.StringVar(); tk.Label(goal_frame, textvariable=self.goal_label_var, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], font=self.controller.primary_font).pack()
        app_list_frame = tk.LabelFrame(self, text="Top Applications", bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], font=self.controller.card_title_font, padx=10, pady=10, borderwidth=0)
        app_list_frame.pack(fill="both", expand=True)
        self.app_list_text = tk.Text(app_list_frame, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], font=self.controller.primary_font, relief="flat", state="disabled", highlightthickness=0)
        self.app_list_text.pack(fill="both", expand=True)
    def create_card(self, parent, title, data_var):
        card = tk.Frame(parent, bg=self.controller.theme_colors["frame"], padx=20, pady=15)
        card.pack(side="left", fill="x", expand=True, padx=10)
        tk.Label(card, text=title, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], font=self.controller.card_title_font).pack(anchor="w")
        tk.Label(card, textvariable=data_var, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["accent"], font=self.controller.card_value_font).pack(anchor="w", pady=(5, 0))
    def on_show(self): self.update_all_stats()
    def update_all_stats(self):
        today_str = date.today().strftime('%Y-%m-%d')
        todays_logs = [entry for entry in self.controller.data if entry.get('time', '').startswith(today_str)]
        active_s, idle_s, app_durations, cat_durations = self.controller.active_time_seconds, self.controller.idle_time_seconds, defaultdict(float), defaultdict(float)
        for i in range(len(todays_logs) - 1):
            try:
                t1 = datetime.strptime(todays_logs[i]['time'], '%Y-%m-%d %H:%M:%S')
                t2 = datetime.strptime(todays_logs[i+1]['time'], '%Y-%m-%d %H:%M:%S')
                duration = (t2 - t1).total_seconds()
                if duration > 1800: continue
                if "Window:" in todays_logs[i].get('event', ''):
                    app_name = todays_logs[i]['event']
                    app_durations[app_name] += duration
                    category = self.controller.categorize_app(app_name)
                    cat_durations[category] += duration
            except: continue
        self.active_time_var.set(self.format_time(active_s)); self.idle_time_var.set(self.format_time(idle_s))
        if app_durations: self.most_used_var.set(max(app_durations, key=app_durations.get).replace("Window: ", "").split(" - ")[-1].strip())
        total_cat_time = cat_durations.get("Work", 0) + cat_durations.get("Entertainment", 0)
        if total_cat_time > 0: self.productivity_var.set(f"{(cat_durations.get('Work', 0) / total_cat_time) * 100:.0f}%")
        else: self.productivity_var.set("--%")
        work_seconds = cat_durations.get("Work", 0)
        goal_hours = self.controller.config.get('daily_work_goal_hours', 4)
        goal_seconds = goal_hours * 3600
        self.goal_progress['value'] = (work_seconds / goal_seconds) * 100 if goal_seconds > 0 else 0
        self.goal_label_var.set(f"{self.format_time(work_seconds)} / {goal_hours} hours")
        self.app_list_text.config(state="normal"); self.app_list_text.delete(1.0, "end")
        sorted_apps = sorted(app_durations.items(), key=lambda item: item[1], reverse=True)
        for app, seconds in sorted_apps[:10]:
            self.app_list_text.insert("end", f"{app.replace('Window: ', '').split(' - ')[-1].strip():<40} {self.format_time(seconds)}\n")
        self.app_list_text.config(state="disabled")
    def format_time(self, seconds):
        h = int(seconds // 3600); m = int((seconds % 3600) // 60)
        return f"{h}h {m}m"
class LogsPage(BasePage):
    def __init__(self, parent, controller):
        super().__init__(parent, controller)
        top_frame = tk.Frame(self, bg=self.controller.theme_colors["bg"]); top_frame.pack(fill="x", pady=10)
        tk.Label(top_frame, text="Select Date:", font=self.controller.header_font, bg=self.controller.theme_colors["bg"], fg=self.controller.theme_colors["text"]).pack(side="left", padx=(0, 10))
        if CALENDAR_ENABLED:
            self.date_picker = DateEntry(top_frame, date_pattern='y-mm-dd', font=self.controller.primary_font)
            self.date_picker.pack(side="left")
            self.date_picker.bind("<<DateEntrySelected>>", self.filter_logs_by_date)
        else: tk.Label(top_frame, text="tkcalendar not installed", font=self.controller.primary_font).pack(side="left")
        notebook = ttk.Notebook(self); notebook.pack(fill="both", expand=True)
        self.activity_tree = self.create_log_tab(notebook, "Activity Log", "ACTIVITY")
        self.click_tree = self.create_log_tab(notebook, "Mouse Clicks", "CLICK EVENT")
        self.clipboard_tree = self.create_log_tab(notebook, "Clipboard History", "COPIED TEXT")
    def create_log_tab(self, notebook, tab_name, column_name):
        tab_frame = tk.Frame(notebook); notebook.add(tab_frame, text=tab_name)
        tree = ttk.Treeview(tab_frame, columns=("Time", "Activity"), show="headings")
        tree.heading("Time", text="TIME"); tree.heading("Activity", text=column_name)
        tree.column("Time", width=180, anchor="w"); tree.column("Activity", width=600, anchor="w")
        scrollbar = ttk.Scrollbar(tab_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y"); tree.pack(side="left", fill="both", expand=True)
        return tree
    def on_show(self): self.filter_logs_by_date()
    def filter_logs_by_date(self, event=None):
        if not CALENDAR_ENABLED: return
        selected_date = self.date_picker.get_date().strftime('%Y-%m-%d')
        for tree in [self.activity_tree, self.click_tree, self.clipboard_tree]:
            if tree: tree.delete(*tree.get_children())
        for entry in self.controller.data:
            if entry.get('time', '').startswith(selected_date):
                self.add_log(entry.get('type', 'activity'), entry.get('time'), entry.get('event'))
    def add_log(self, event_type, timestamp, event_description):
        target_tree = {'click': self.click_tree, 'clipboard': self.clipboard_tree}.get(event_type, self.activity_tree)
        if target_tree and target_tree.winfo_exists():
            target_tree.insert("", "end", values=(timestamp, event_description)); target_tree.yview_moveto(1)
class ReportsPage(BasePage):
    def __init__(self, parent, controller):
        super().__init__(parent, controller)
        if not MATPLOTLIB_ENABLED:
            tk.Label(self, text="matplotlib not installed. Reports are unavailable.", font=self.controller.header_font, bg=self.controller.theme_colors["bg"], fg="red").pack(pady=50)
            return
        left_frame = tk.Frame(self, bg=self.controller.theme_colors["bg"]); left_frame.pack(side="left", fill="y", padx=10)
        tk.Label(left_frame, text="Available Dates", font=self.controller.header_font, bg=self.controller.theme_colors["bg"], fg=self.controller.theme_colors["text"]).pack(pady=5)
        self.date_listbox = tk.Listbox(left_frame, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], selectbackground=self.controller.theme_colors["accent"], relief="flat", exportselection=False)
        self.date_listbox.pack(fill="y", expand=True)
        self.date_listbox.bind("<<ListboxSelect>>", self.generate_report)
        self.chart_frame = tk.Frame(self, bg=self.controller.theme_colors["bg"]); self.chart_frame.pack(side="left", fill="both", expand=True)
    def on_show(self):
        self.populate_dates()
        if self.date_listbox.size() > 0:
            self.date_listbox.selection_set(0)
            self.generate_report()
    def populate_dates(self):
        self.date_listbox.delete(0, tk.END)
        unique_dates = sorted(list(set(datetime.strptime(e['time'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d') for e in self.controller.data)), reverse=True)
        for d in unique_dates: self.date_listbox.insert(tk.END, d)
    def generate_report(self, event=None):
        if not MATPLOTLIB_ENABLED or not self.date_listbox.curselection(): return
        for widget in self.chart_frame.winfo_children(): widget.destroy()
        selected_date_str = self.date_listbox.get(self.date_listbox.curselection())
        filtered_logs = [e for e in self.controller.data if e['time'].startswith(selected_date_str)]
        category_durations = defaultdict(float)
        for i in range(len(filtered_logs) - 1):
            try:
                t1 = datetime.strptime(filtered_logs[i]['time'], '%Y-%m-%d %H:%M:%S')
                t2 = datetime.strptime(filtered_logs[i+1]['time'], '%Y-%m-%d %H:%M:%S')
                duration = (t2 - t1).total_seconds()
                if duration > 3600: continue
                if "Window:" in filtered_logs[i].get('event', ''):
                    category = self.controller.categorize_app(filtered_logs[i]['event'])
                    category_durations[category] += duration
            except: continue
        if not category_durations:
            tk.Label(self.chart_frame, text="No activity data for this day.", font=self.controller.header_font, bg=self.controller.theme_colors["bg"], fg=self.controller.theme_colors["text"]).pack(pady=50)
            return
        fig = Figure(figsize=(5, 4), dpi=100, facecolor=self.controller.theme_colors["bg"])
        ax = fig.add_subplot(111)
        labels = category_durations.keys(); sizes = [v / 3600 for v in category_durations.values()]
        ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, textprops={'color': self.controller.theme_colors["text"]})
        ax.axis('equal'); ax.set_title(f"Time Distribution for {selected_date_str}", color=self.controller.theme_colors["text"])
        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame); canvas.draw(); canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
class SettingsPage(BasePage):
    def __init__(self, parent, controller):
        super().__init__(parent, controller)
        self.scrollable_frame = tk.Frame(self, bg=self.controller.theme_colors["bg"])
        self.scrollable_frame.pack(fill="both", expand=True)
        # --- Monitoring Settings ---
        monitor_frame = self.create_settings_group("Monitoring Permissions")
        self.clicks_var = tk.BooleanVar(value=self.controller.config.get('monitor_clicks', True))
        self.create_checkbutton(monitor_frame, "Monitor Mouse Clicks", self.clicks_var)
        self.clipboard_var = tk.BooleanVar(value=self.controller.config.get('monitor_clipboard', True))
        self.create_checkbutton(monitor_frame, "Monitor Clipboard", self.clipboard_var)
        # --- General Settings ---
        general_frame = self.create_settings_group("General Settings")
        self.create_spinbox(general_frame, "Idle Threshold (minutes):", 1, 60, self.controller.config.get('idle_threshold_minutes', 5))
        self.create_theme_selector(general_frame, "Theme:")
        # --- Goal Settings ---
        goal_frame = self.create_settings_group("Daily Goal")
        self.create_spinbox(goal_frame, "Work Goal (hours):", 1, 12, self.controller.config.get('daily_work_goal_hours', 4))
        # --- Break Reminder ---
        break_frame = self.create_settings_group("Break Reminder")
        self.break_var = tk.BooleanVar(value=self.controller.config.get('break_reminder_enabled', True))
        self.create_checkbutton(break_frame, "Enable Break Reminders", self.break_var)
        self.create_spinbox(break_frame, "Remind after (minutes):", 15, 120, self.controller.config.get('break_reminder_interval_minutes', 60))
        # --- Category Settings ---
        category_frame = self.create_settings_group("App Categories (comma-separated keywords)")
        self.work_keywords_entry = self.create_text_entry(category_frame, "Work Keywords:", self.controller.config.get('work_keywords', ''))
        self.entertainment_keywords_entry = self.create_text_entry(category_frame, "Entertainment Keywords:", self.controller.config.get('entertainment_keywords', ''))
        # --- Data Export ---
        data_frame = self.create_settings_group("Data Management")
        tk.Button(data_frame, text="Export All Data to CSV", command=self.export_to_csv, font=self.controller.button_font, bg=self.controller.theme_colors["accent"], fg="#FFFFFF", relief="flat").pack(pady=5)
        # --- Save Button ---
        save_button = tk.Button(self, text="Save Settings", command=self.save_settings, font=self.controller.button_font, bg=self.controller.theme_colors["accent"], fg="#FFFFFF", relief="flat")
        save_button.pack(pady=20)
    def create_settings_group(self, text):
        frame = tk.LabelFrame(self.scrollable_frame, text=text, font=self.controller.header_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], padx=15, pady=10, borderwidth=0)
        frame.pack(fill="x", padx=20, pady=10)
        return frame
    def create_checkbutton(self, parent, text, variable):
        frame = tk.Frame(parent, bg=self.controller.theme_colors["frame"])
        frame.pack(fill="x", pady=2)
        style = ttk.Style()
        style.configure("Switch.TCheckbutton", background=self.controller.theme_colors["frame"], foreground=self.controller.theme_colors["text"])
        cb = ttk.Checkbutton(frame, text=text, variable=variable, style="Switch.TCheckbutton")
        cb.pack(side="left")
    def create_spinbox(self, parent, text, from_, to, default):
        frame = tk.Frame(parent, bg=self.controller.theme_colors["frame"]); frame.pack(fill="x", pady=5)
        tk.Label(frame, text=text, font=self.controller.primary_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"]).pack(side="left")
        spinbox = ttk.Spinbox(frame, from_=from_, to=to, width=5, font=self.controller.primary_font); spinbox.set(default)
        spinbox.pack(side="left", padx=10)
        setattr(self, text.lower().replace(" ", "_").replace(":", "").replace("(", "").replace(")", ""), spinbox)
    def create_theme_selector(self, parent, text):
        frame = tk.Frame(parent, bg=self.controller.theme_colors["frame"]); frame.pack(fill="x", pady=5)
        tk.Label(frame, text=text, font=self.controller.primary_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"]).pack(side="left")
        self.theme_var = tk.StringVar(value=self.controller.config.get('theme', 'WordPress Light'))
        themes = ["WordPress Light", "Dark Teal", "Light Mint"]
        style = ttk.Style()
        style.configure("TMenubutton", background=self.controller.theme_colors["frame"], foreground=self.controller.theme_colors["text"])
        option_menu = ttk.OptionMenu(frame, self.theme_var, self.controller.config.get('theme', 'WordPress Light'), *themes)
        option_menu.pack(side="left", padx=10)
    def create_text_entry(self, parent, text, default):
        frame = tk.Frame(parent, bg=self.controller.theme_colors["frame"]); frame.pack(fill="x", pady=5)
        tk.Label(frame, text=text, font=self.controller.primary_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"]).pack(side="left", anchor="n")
        entry = tk.Text(frame, height=2, width=60, font=self.controller.primary_font, bg="#DDDDDD" if self.controller.config.get('theme') == 'WordPress Light' else "#333333", fg=self.controller.theme_colors["text"], relief="solid", borderwidth=1, insertbackground=self.controller.theme_colors["text"])
        entry.insert("1.0", default)
        entry.pack(side="left", padx=10, expand=True, fill="x")
        return entry
    def save_settings(self):
        self.controller.config['monitor_clicks'] = self.clicks_var.get()
        self.controller.config['monitor_clipboard'] = self.clipboard_var.get()
        self.controller.config['idle_threshold_minutes'] = int(self.idle_threshold_minutes_.get())
        self.controller.config['theme'] = self.theme_var.get()
        self.controller.config['daily_work_goal_hours'] = int(self.work_goal_hours_.get())
        self.controller.config['break_reminder_enabled'] = self.break_var.get()
        self.controller.config['break_reminder_interval_minutes'] = int(self.remind_after_minutes_.get())
        self.controller.config['work_keywords'] = self.work_keywords_entry.get("1.0", "end-1c")
        self.controller.config['entertainment_keywords'] = self.entertainment_keywords_entry.get("1.0", "end-1c")
        self.controller.save_config()
        messagebox.showinfo("Settings Saved", "Your settings have been saved. Please restart the application for all changes to take effect.")
    def export_to_csv(self):
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")], initialfile="activity_log_export.csv")
            if not file_path: return
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Timestamp", "Type", "Event"])
                for entry in self.controller.data:
                    writer.writerow([entry.get('time'), entry.get('type'), entry.get('event')])
            messagebox.showinfo("Export Successful", f"Data successfully exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"An error occurred during export: {e}")
    def on_show(self): pass

# --- NEW: About Page ---
class AboutPage(BasePage):
    def __init__(self, parent, controller):
        super().__init__(parent, controller)
        
        container = tk.Frame(self, bg=self.controller.theme_colors["frame"])
        container.pack(expand=True)

        tk.Label(container, text="Developed By", font=self.controller.header_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"]).pack(pady=(20, 5))
        tk.Label(container, text="Muhammad Al-amin", font=self.controller.card_value_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["accent"]).pack()

        tk.Label(container, text="Contact Information", font=self.controller.header_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"]).pack(pady=(30, 5))
        
        self.create_info_row(container, "Phone:", "+8801778189644")
        self.create_link_row(container, "Email:", "mdalaminkhalifa2002@gmail.com", "mailto:mdalaminkhalifa2002@gmail.com")
        self.create_link_row(container, "Facebook:", "https://www.facebook.com/mdalamins20", "https://www.facebook.com/mdalamins20")

    def create_info_row(self, parent, title, value):
        frame = tk.Frame(parent, bg=self.controller.theme_colors["frame"])
        frame.pack(pady=5, padx=20, fill='x')
        tk.Label(frame, text=title, font=self.controller.primary_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], width=10, anchor='w').pack(side="left")
        tk.Label(frame, text=value, font=self.controller.primary_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"]).pack(side="left")

    def create_link_row(self, parent, title, text, url):
        frame = tk.Frame(parent, bg=self.controller.theme_colors["frame"])
        frame.pack(pady=5, padx=20, fill='x')
        tk.Label(frame, text=title, font=self.controller.primary_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["text"], width=10, anchor='w').pack(side="left")
        link = tk.Label(frame, text=text, font=self.controller.link_font, bg=self.controller.theme_colors["frame"], fg=self.controller.theme_colors["accent"], cursor="hand2")
        link.pack(side="left")
        link.bind("<Button-1>", lambda e: webbrowser.open_new(url))

if __name__ == '__main__':
    if IDLE_DETECTION_ENABLED:
        threading.Thread(target=start_listeners, daemon=True).start()
    root = tk.Tk()
    app = ActivityLoggerApp(root)
    root.mainloop()


